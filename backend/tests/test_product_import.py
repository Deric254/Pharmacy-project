"""
Product bulk import tests. The properties that matter:
  1. The template is a real, valid spreadsheet with real data
     validation rules (dropdown for unit, numeric-only for quantity
     and price) -- not just headers.
  2. Import is genuinely all-or-nothing: if ANY row has a problem,
     ZERO rows are imported, even the clean ones in the same file.
  3. Every problem is reported at once, with the real row number, not
     just the first one found.
  4. Duplicates are caught both within the uploaded file itself and
     against the existing catalog -- case-insensitively for names.
  5. The template's own EXAMPLE row is silently skipped if left in,
     never treated as a real product or an error.
"""

import io

import openpyxl

from app.services.product_import_service import generate_import_template


async def _login(client, username: str, password: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return str(r.json()["access_token"])


def _build_workbook(rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Barcode", "Unit", "Reorder point", "Selling price"])
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestImportTemplate:
    def test_template_is_a_real_spreadsheet_with_real_validation(self):
        content = generate_import_template()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        assert [c.value for c in ws[1]][:5] == [
            "Name",
            "Barcode",
            "Unit",
            "Reorder point",
            "Selling price",
        ]
        # Real Excel data validation, not just header text -- this is
        # what makes a typo'd unit or a non-numeric price structurally
        # hard to enter in the first place.
        validations = list(ws.data_validations.dataValidation)
        assert len(validations) == 3
        types = {v.type for v in validations}
        assert types == {"list", "whole", "decimal"}

    async def test_template_download_requires_permission(self, client, employee_user):
        token = await _login(client, "joe", "pass1234")
        r = await client.get(
            "/api/v1/products/import-template", headers={"Authorization": f"Bearer {token}"}
        )
        assert r.status_code == 403


class TestBulkImport:
    async def test_clean_file_imports_everything(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["Amoxicillin 500mg", "AMX500", "capsule", 15, 25.0],
                ["Paracetamol 500mg", "", "tablet", 20, 10.0],
                ["Ibuprofen 200mg", "IBU200", "tablet", 10, 12.5],
            ]
        )
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 201
        assert r.json()["created"] == 3

        products = await client.get(
            "/api/v1/products", headers={"Authorization": f"Bearer {token}"}
        )
        names = {p["name"] for p in products.json()}
        assert names == {"Amoxicillin 500mg", "Paracetamol 500mg", "Ibuprofen 200mg"}

    async def test_one_bad_row_rejects_the_whole_file_not_just_that_row(self, client, owner_user):
        """
        The actual guarantee this proves: a file with one genuinely
        clean row and one genuinely broken row must import ZERO rows,
        not one. A partial import is exactly the inconsistent, silent-
        corruption risk this whole feature exists to prevent.
        """
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["Genuinely Clean Row", "", "tablet", 10, 8.0],
                ["Bad Reorder Point", "", "tablet", "not-a-number", 5.0],
            ]
        )
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        errors = r.json()["detail"]["errors"]
        assert len(errors) == 1
        assert errors[0]["field"] == "Reorder point"

        products = await client.get(
            "/api/v1/products", headers={"Authorization": f"Bearer {token}"}
        )
        # The clean row must NOT have snuck in alongside the rejected one.
        assert products.json() == []

    async def test_every_problem_reported_at_once_not_just_the_first(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["Bad Unit Row", "", "not-a-real-unit", 10, 5.0],
                ["Bad Price Row", "", "tablet", 10, "not-a-number"],
                ["Bad Reorder Row", "", "tablet", "not-a-number", 5.0],
            ]
        )
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        assert len(r.json()["detail"]["errors"]) == 3

    async def test_within_file_duplicate_name_is_caught_case_insensitively(
        self, client, owner_user
    ):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["Metformin 500mg", "", "tablet", 10, 9.0],
                ["metformin 500mg", "", "tablet", 10, 9.0],
            ]
        )
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        errors = r.json()["detail"]["errors"]
        assert len(errors) == 1
        assert "Duplicate of row" in errors[0]["message"]

    async def test_duplicate_against_existing_catalog_is_caught(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}
        await client.post("/api/v1/products", json={"name": "Already Exists"}, headers=headers)

        content = _build_workbook([["Already Exists", "", "tablet", 10, 5.0]])
        r = await client.post(
            "/api/v1/products/import",
            headers=headers,
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        assert "already exists in the catalog" in r.json()["detail"]["errors"][0]["message"]

    async def test_example_row_is_silently_skipped_not_imported_or_flagged(
        self, client, owner_user
    ):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["EXAMPLE - Paracetamol 500mg", "", "tablet", 20, 15.0],
                ["Real Product", "", "tablet", 10, 5.0],
            ]
        )
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 201
        assert r.json()["created"] == 1  # only the real row, example silently excluded

        products = await client.get(
            "/api/v1/products", headers={"Authorization": f"Bearer {token}"}
        )
        names = {p["name"] for p in products.json()}
        assert names == {"Real Product"}

    async def test_import_requires_permission(self, client, employee_user):
        token = await _login(client, "joe", "pass1234")
        content = _build_workbook([["Some Product", "", "tablet", 10, 5.0]])
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 403

    async def test_non_excel_file_rejected_cleanly(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("notes.txt", b"this is not a spreadsheet", "text/plain")},
        )
        assert r.status_code == 400

    async def test_empty_file_with_no_rows_rejected_cleanly(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook([])
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422

    async def test_name_over_150_chars_is_a_clean_error_not_a_crash(self, client, owner_user):
        """
        The real bug an adversarial chaos test found: a name longer
        than the schema allows crashed with an unhandled 500 instead
        of being reported like every other bad-row problem. Must come
        back as a normal, structured per-row error.
        """
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook([["A" * 500, "", "tablet", 10, 5.0]])
        r = await client.post(
            "/api/v1/products/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        errors = r.json()["detail"]["errors"]
        assert any("150 characters" in e["message"] for e in errors)
