"""
Purchase order bulk import tests -- the direct path. The properties
that matter:
  1. The template is real, with real numeric validation.
  2. Product names match the real, active catalog case-insensitively
     -- never invents a product from an unmatched name.
  3. Import is genuinely all-or-nothing: any problem anywhere in the
     file means nothing is received at all.
  4. A successful import goes straight to a RECEIVED purchase order
     with real batches and real stock, matching quick_purchase()
     exactly -- no draft/send/in-transit ceremony.
"""

import io

import openpyxl

from app.core.database import AsyncSessionLocal
from app.models.product import Product
from app.models.supplier import Supplier
from app.services.purchase_order_import_service import generate_purchase_order_import_template


async def _login(client, username: str, password: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return str(r.json()["access_token"])


async def _make_supplier(name: str = "PO Import Supplier") -> int:
    async with AsyncSessionLocal() as db:
        supplier = Supplier(name=name)
        db.add(supplier)
        await db.commit()
        return int(supplier.id)


async def _make_product(name: str, price: float = 20.0) -> int:
    async with AsyncSessionLocal() as db:
        product = Product(name=name, default_selling_price=price)
        db.add(product)
        await db.commit()
        return int(product.id)


def _build_workbook(rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product name", "Quantity", "Batch number", "Expiry date", "Unit cost"])
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestPurchaseOrderImportTemplate:
    def test_template_is_real_with_real_validation(self):
        content = generate_purchase_order_import_template()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        assert [c.value for c in ws[1]][:5] == [
            "Product name",
            "Quantity",
            "Batch number",
            "Expiry date",
            "Unit cost",
        ]
        validations = list(ws.data_validations.dataValidation)
        assert len(validations) == 2
        assert {v.type for v in validations} == {"whole", "decimal"}


class TestPurchaseOrderBulkImport:
    async def test_clean_file_goes_straight_to_received_with_real_stock(self, client, owner_user):
        supplier_id = await _make_supplier()
        await _make_product("Amoxicillin 500mg")
        await _make_product("Paracetamol 500mg")
        token = await _login(client, "lucy", "S3curePass!")

        content = _build_workbook(
            [
                ["Amoxicillin 500mg", 100, "AMX-001", "2027-06-30", 10.0],
                ["paracetamol 500mg", 200, "PARA-001", "2027-06-30", 4.0],  # lowercase on purpose
            ]
        )
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers={"Authorization": f"Bearer {token}"},
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 201
        body = r.json()
        # Straight to RECEIVED -- no draft/send/in-transit ceremony.
        assert body["status"] == "RECEIVED"
        assert body["received_at"] is not None
        names = {item["product_name"] for item in body["items"]}
        assert names == {"Amoxicillin 500mg", "Paracetamol 500mg"}

        products = await client.get(
            "/api/v1/products", headers={"Authorization": f"Bearer {token}"}
        )
        stock = {p["name"]: p["total_qty_available"] for p in products.json()}
        assert stock["Amoxicillin 500mg"] == 100
        assert stock["Paracetamol 500mg"] == 200

    async def test_unmatched_product_name_rejects_the_whole_file(self, client, owner_user):
        supplier_id = await _make_supplier()
        await _make_product("Real Product Here")
        token = await _login(client, "lucy", "S3curePass!")

        content = _build_workbook(
            [
                ["Real Product Here", 10, "B1", "2027-06-30", 5.0],
                ["Does Not Exist In Catalog", 10, "B2", "2027-06-30", 5.0],
            ]
        )
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers={"Authorization": f"Bearer {token}"},
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        errors = r.json()["detail"]["errors"]
        assert any("No active product" in e["message"] for e in errors)

        products = await client.get(
            "/api/v1/products", headers={"Authorization": f"Bearer {token}"}
        )
        # Not even the matched row's stock was received.
        stock = next(p for p in products.json() if p["name"] == "Real Product Here")[
            "total_qty_available"
        ]
        assert stock == 0

    async def test_negative_quantity_and_duplicate_batch_both_reported_at_once(
        self, client, owner_user
    ):
        supplier_id = await _make_supplier()
        await _make_product("Duplicate Batch Test Product")
        token = await _login(client, "lucy", "S3curePass!")

        content = _build_workbook(
            [
                ["Duplicate Batch Test Product", -5, "SAMEBATCH", "2027-06-30", 5.0],
                ["Duplicate Batch Test Product", 10, "SAMEBATCH", "2027-06-30", 5.0],
            ]
        )
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers={"Authorization": f"Bearer {token}"},
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        errors = r.json()["detail"]["errors"]
        assert any("whole number" in e["message"] for e in errors)
        assert any("Duplicate of row" in e["message"] for e in errors)

    async def test_invalid_expiry_date_rejected_cleanly(self, client, owner_user):
        supplier_id = await _make_supplier()
        await _make_product("Bad Expiry Test Product")
        token = await _login(client, "lucy", "S3curePass!")

        content = _build_workbook([["Bad Expiry Test Product", 10, "B1", "not-a-real-date", 5.0]])
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers={"Authorization": f"Bearer {token}"},
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        assert any("date" in e["message"].lower() for e in r.json()["detail"]["errors"])

    async def test_deactivated_product_is_treated_as_unmatched(self, client, owner_user):
        supplier_id = await _make_supplier()
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create = await client.post(
            "/api/v1/products", json={"name": "Soon Deactivated"}, headers=headers
        )
        product_id = create.json()["id"]
        await client.delete(f"/api/v1/products/{product_id}", headers=headers)

        content = _build_workbook([["Soon Deactivated", 10, "B1", "2027-06-30", 5.0]])
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers=headers,
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422

    async def test_nonexistent_supplier_rejected_cleanly(self, client, owner_user):
        await _make_product("Some Product")
        token = await _login(client, "lucy", "S3curePass!")

        content = _build_workbook([["Some Product", 10, "B1", "2027-06-30", 5.0]])
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers={"Authorization": f"Bearer {token}"},
            data={"supplier_id": "999999"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 404

    async def test_requires_create_po_permission(self, client, employee_user):
        token = await _login(client, "joe", "pass1234")
        content = _build_workbook([["Some Product", 10, "B1", "2027-06-30", 5.0]])
        r = await client.post(
            "/api/v1/purchase-orders/import",
            headers={"Authorization": f"Bearer {token}"},
            data={"supplier_id": "1"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 403

    async def test_uploading_the_same_list_twice_merges_not_duplicates(self, client, owner_user):
        """
        The real, confirmed bug this closes: uploading the same
        purchase list a second time (or an updated one for the same
        real-world delivery) created a second, separate batch row
        instead of adding to the one already there -- genuine
        inventory duplication despite being the same product and the
        same batch number. This is the exact real-world scenario:
        upload once, upload again, confirm exactly one batch exists
        with the combined quantity, not two.
        """
        product_id = await _make_product("Reupload Test Product")
        supplier_id = await _make_supplier()
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        content = _build_workbook([["Reupload Test Product", 100, "REUP1", "2027-06-30", 10.0]])

        first = await client.post(
            "/api/v1/purchase-orders/import",
            headers=headers,
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert first.status_code == 201

        second = await client.post(
            "/api/v1/purchase-orders/import",
            headers=headers,
            data={"supplier_id": str(supplier_id)},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert second.status_code == 201

        batches = await client.get(f"/api/v1/products/{product_id}/batches", headers=headers)
        matching = [b for b in batches.json() if b["batch_number"] == "REUP1"]
        assert len(matching) == 1, "Expected exactly one batch, found duplication"
        assert matching[0]["qty_remaining"] == 200
