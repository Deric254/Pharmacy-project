"""
Customer bulk import. Same shape and same guarantees as product bulk
import: the template constrains what can be typed in, but the real
authoritative guarantee is server-side -- every row fully validated,
and if anything is wrong, nothing is imported at all.
"""

import io
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.customer import Customer
from app.schemas.customer import CustomerCreate
from app.schemas.product import BulkImportResult, ImportRowError

_HEADERS = ["Name", "Phone", "Email"]
_EXAMPLE_ROW = ["EXAMPLE - Jane Mwangi", "0712345678", "jane@example.com"]
_MAX_ROWS = 2000


def generate_customer_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Customers"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    example_font = Font(name="Arial", italic=True, color="6B7280")
    for col, value in enumerate(_EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = example_font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    instructions = ws.cell(row=1, column=5, value="Delete the EXAMPLE row before importing.")
    instructions.font = Font(name="Arial", italic=True, size=9, color="991B1B")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _clean_str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


async def _parse_and_validate(
    db: AsyncSession, file_bytes: bytes
) -> tuple[list[CustomerCreate], list[ImportRowError]]:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="This file has no worksheet to read.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail="Could not read this file as an Excel spreadsheet."
        ) from exc

    errors: list[ImportRowError] = []
    candidates: list[CustomerCreate] = []
    seen_names: dict[str, int] = {}
    seen_phones: dict[str, int] = {}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(rows) > _MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"This file has more than {_MAX_ROWS} rows. Split it into smaller batches.",
        )

    for offset, row in enumerate(rows):
        row_num = offset + 2
        row_values: list[Any] = (list(row) + [None] * 3)[:3]
        name_raw, phone_raw, email_raw = row_values
        name = _clean_str(name_raw)

        if not name:
            continue
        if name.upper().startswith("EXAMPLE"):
            continue

        phone = _clean_str(phone_raw) or None
        email = _clean_str(email_raw) or None

        row_already_invalid = False
        if len(name) > 150:
            errors.append(
                ImportRowError(
                    row=row_num, field="Name", message="Must be 150 characters or fewer."
                )
            )
            row_already_invalid = True
        if phone and len(phone) > 30:
            errors.append(
                ImportRowError(
                    row=row_num, field="Phone", message="Must be 30 characters or fewer."
                )
            )
            row_already_invalid = True
        if email and len(email) > 120:
            errors.append(
                ImportRowError(
                    row=row_num, field="Email", message="Must be 120 characters or fewer."
                )
            )
            row_already_invalid = True

        name_key = name.lower()
        if name_key in seen_names:
            errors.append(
                ImportRowError(
                    row=row_num,
                    field="Name",
                    message=f"Duplicate of row {seen_names[name_key]} in this same file.",
                )
            )
        else:
            seen_names[name_key] = row_num

        if phone:
            if phone in seen_phones:
                errors.append(
                    ImportRowError(
                        row=row_num,
                        field="Phone",
                        message=f"Duplicate of row {seen_phones[phone]} in this same file.",
                    )
                )
            else:
                seen_phones[phone] = row_num

        if row_already_invalid:
            continue

        try:
            candidates.append(CustomerCreate(name=name, phone=phone, email=email))
        except ValidationError as exc:
            errors.append(
                ImportRowError(
                    row=row_num, field="Row", message=f"Invalid data: {exc.errors()[0]['msg']}"
                )
            )

    if not candidates:
        errors.append(
            ImportRowError(row=0, field="File", message="No customer rows found in this file.")
        )
        return candidates, errors

    phones_to_check = [c.phone for c in candidates if c.phone]
    existing_phone_set: set[str] = set()
    if phones_to_check:
        existing_phones = await db.execute(
            select(Customer.phone).where(Customer.phone.in_(phones_to_check))
        )
        existing_phone_set = {row[0] for row in existing_phones.all()}

    for candidate in candidates:
        if candidate.phone and candidate.phone in existing_phone_set:
            errors.append(
                ImportRowError(
                    row=seen_phones[candidate.phone],
                    field="Phone",
                    message=f'Phone "{candidate.phone}" already belongs to an existing customer.',
                )
            )

    return candidates, errors


async def bulk_import_customers(db: AsyncSession, file_bytes: bytes) -> BulkImportResult:
    candidates, errors = await _parse_and_validate(db, file_bytes)

    if errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": f"{len(errors)} problem(s) found. Nothing was imported.",
                "errors": [e.model_dump() for e in errors],
            },
        )

    for candidate in candidates:
        db.add(Customer(**candidate.model_dump()))
    await db.commit()

    return BulkImportResult(created=len(candidates))
