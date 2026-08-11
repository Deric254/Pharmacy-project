import io
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.concurrency import run_in_threadpool

from app.models.stock_take import StockTake, StockTakeItem, StockTakeStatus
from app.models.user import User
from app.schemas.inventory import AdjustmentReason
from app.schemas.stock_take import CountSubmit, StockTakeOut
from app.services.stock_take_service import StockTakeService

_HEADERS = ["Product name", "Batch number", "Expiry date", "System quantity", "Physical quantity"]
_ID_COLUMN = 6  # hidden -- carries the real item id for exact matching, never shown or typed


async def _load_open_stock_take(db: AsyncSession, stock_take_id: int) -> StockTake:
    result = await db.execute(select(StockTake).where(StockTake.id == stock_take_id))
    stock_take = result.scalar_one_or_none()
    if stock_take is None:
        raise HTTPException(status_code=404, detail="Stock take not found")
    if stock_take.status != StockTakeStatus.OPEN:
        raise HTTPException(status_code=400, detail="Stock take is already closed")
    return stock_take


async def generate_count_template(db: AsyncSession, stock_take_id: int) -> bytes:
    await _load_open_stock_take(db, stock_take_id)

    items_result = await db.execute(
        select(StockTakeItem)
        .where(StockTakeItem.stock_take_id == stock_take_id)
        .order_by(StockTakeItem.id)
    )
    items = items_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Stock Count"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws.cell(row=1, column=_ID_COLUMN, value="Item ID")
    ws.freeze_panes = "A2"

    for row_num, item in enumerate(items, start=2):
        ws.cell(row=row_num, column=1, value=item.product_name)
        ws.cell(row=row_num, column=2, value=item.batch_number)
        ws.cell(row=row_num, column=3, value=item.batch.expiry_date.isoformat())
        ws.cell(row=row_num, column=4, value=item.expected_qty)
        # Physical quantity (column 5) left blank -- the person fills
        # this in with what they actually counted.
        ws.cell(row=row_num, column=_ID_COLUMN, value=item.id)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    # Hidden -- this is bookkeeping for the re-upload, not something
    # anyone needs to see or, worse, accidentally edit.
    ws.column_dimensions["F"].hidden = True

    if items:
        qty_validation = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1=0,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid quantity",
            error="Physical quantity must be a whole number, 0 or more.",
        )
        ws.add_data_validation(qty_validation)
        qty_validation.add(f"E2:E{len(items) + 1}")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def import_counts(
    db: AsyncSession, stock_take_id: int, file_bytes: bytes, user: User
) -> StockTakeOut:
    await _load_open_stock_take(db, stock_take_id)

    try:
        wb = await run_in_threadpool(load_workbook, io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="This file has no worksheet to read.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail="Could not read this file as an Excel spreadsheet."
        ) from exc

    # Every real item for this stock take, keyed by id -- used both to
    # validate every row up front (all-or-nothing, same as every other
    # bulk import in this system) and to apply counts afterward.
    items_result = await db.execute(
        select(StockTakeItem).where(StockTakeItem.stock_take_id == stock_take_id)
    )
    items_by_id = {item.id: item for item in items_result.scalars().all()}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    counts_to_apply: list[tuple[int, int]] = []  # (item_id, physical_qty)
    seen_item_ids: set[int] = set()

    for offset, row in enumerate(rows):
        row_num = offset + 2
        row_values: list[Any] = (list(row) + [None] * 6)[:6]
        _, _, _, _, physical_raw, item_id_raw = row_values

        if item_id_raw is None:
            continue  # a genuinely blank trailing row -- not an error

        try:
            item_id = int(item_id_raw)
        except (TypeError, ValueError) as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Row {row_num}: the hidden Item ID column was edited or is missing. "
                "Please use the downloaded template unmodified.",
            ) from exc

        item = items_by_id.get(item_id)
        if item is None:
            raise HTTPException(
                status_code=400,
                detail=f"Row {row_num}: this item does not belong to this stock take.",
            )
        if item_id in seen_item_ids:
            raise HTTPException(
                status_code=400, detail=f"Row {row_num}: duplicate row for the same item."
            )
        seen_item_ids.add(item_id)

        if physical_raw is None or physical_raw == "":
            continue  # not yet counted -- leave it for later, don't fail the whole import

        try:
            physical_qty = int(physical_raw)
            if physical_qty < 0:
                raise ValueError
        except (TypeError, ValueError) as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Row {row_num}: Physical quantity must be a whole number, 0 or more.",
            ) from exc

        counts_to_apply.append((item_id, physical_qty))

    if not counts_to_apply:
        raise HTTPException(
            status_code=422, detail="No physical quantities were filled in this file."
        )

    # Every row already validated above -- applying now reuses the
    # exact same per-item logic as counting by hand (same self-approve
    # threshold, same concurrency-safe approval claim), so a bulk
    # import behaves identically to doing it one at a time, just faster.
    service = StockTakeService(db)
    for item_id, physical_qty in counts_to_apply:
        item = items_by_id[item_id]
        variance = physical_qty - item.expected_qty
        reason = AdjustmentReason.MISCOUNT if variance != 0 else None
        await service.submit_count(
            stock_take_id, item_id, CountSubmit(physical_qty=physical_qty, reason=reason), user
        )

    # If every item is now counted and resolved, this finishes the
    # whole stock take in the same step -- if anything still needs a
    # manager's approval (a large variance), close() correctly leaves
    # it open rather than forcing it through.
    try:
        return await service.close(stock_take_id, user)
    except HTTPException:
        return await service.get(stock_take_id)
