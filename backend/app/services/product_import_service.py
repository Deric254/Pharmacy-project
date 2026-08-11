"""
Product bulk import.

Two layers of defense, not one. The Excel template itself constrains
what can be typed into it (a dropdown for unit, numeric-only cells for
quantity and price) -- that's what makes a bad row hard to create in
the first place. But Excel-level validation can be bypassed (pasting
values, editing with a different tool, a formula that evaluates past
the constraint), so it is never trusted as the real guarantee. Every
row is fully re-validated here, server-side, and if ANY row has ANY
problem, NOTHING is imported -- no partial import, no "47 succeeded, 3
failed" leaving the catalog in a half-imported state. Either the whole
file is clean and all of it lands in one transaction, or none of it
does, and the response says exactly which rows and fields need fixing
so the file can be corrected and re-uploaded whole.
"""

import io
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.concurrency import run_in_threadpool

from app.models.product import Product
from app.schemas.product import BulkImportResult, ImportRowError, ProductCreate

_COMMON_UNITS = [
    "unit",
    "tablet",
    "capsule",
    "bottle",
    "box",
    "syrup",
    "injection",
    "vial",
    "tube",
    "sachet",
    "strip",
    "pack",
]

_HEADERS = ["Name", "Barcode", "Unit", "Reorder point", "Selling price"]
_EXAMPLE_ROW: list[str | int | float] = ["EXAMPLE - Paracetamol 500mg", "", "tablet", 20, 15.0]
_MAX_ROWS = 2000  # generous for a small pharmacy's catalog; guards against an accidental huge file


def generate_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # a freshly created Workbook always has an active sheet
    ws.title = "Products"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    example_font = Font(name="Arial", italic=True, color="6B7280")
    for col, value in enumerate(_EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = example_font

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # Unit: a dropdown, not free text -- the actual mechanism that makes
    # "Tabs" / "tabs " / "Tablet " typo variants structurally impossible
    # instead of merely discouraged.
    unit_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(_COMMON_UNITS)}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid unit",
        error="Choose a unit from the dropdown list.",
    )
    ws.add_data_validation(unit_validation)
    unit_validation.add(f"C2:C{_MAX_ROWS}")

    reorder_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=0,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid reorder point",
        error="Reorder point must be a whole number, 0 or greater.",
    )
    ws.add_data_validation(reorder_validation)
    reorder_validation.add(f"D2:D{_MAX_ROWS}")

    price_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1=0,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid selling price",
        error="Selling price must be a number, 0 or greater.",
    )
    ws.add_data_validation(price_validation)
    price_validation.add(f"E2:E{_MAX_ROWS}")

    instructions = ws.cell(row=1, column=7, value="Delete the EXAMPLE row before importing.")
    instructions.font = Font(name="Arial", italic=True, size=9, color="991B1B")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _clean_str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


async def _parse_and_validate(
    db: AsyncSession, file_bytes: bytes
) -> tuple[list[ProductCreate], list[ImportRowError]]:
    try:
        wb = await run_in_threadpool(load_workbook, io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="This file has no worksheet to read.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail="Could not read this file as an Excel spreadsheet."
        ) from exc

    errors: list[ImportRowError] = []
    candidates: list[ProductCreate] = []
    seen_names: dict[str, int] = {}  # lowercased name -> first row it appeared on
    seen_barcodes: dict[str, int] = {}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(rows) > _MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"This file has more than {_MAX_ROWS} rows. Split it into smaller batches.",
        )

    for offset, row in enumerate(rows):
        row_num = offset + 2  # 1-indexed, header is row 1
        row_values: list[Any] = (list(row) + [None] * 5)[:5]
        name_raw, barcode_raw, unit_raw, reorder_raw, price_raw = row_values
        name = _clean_str(name_raw)

        if not name:
            continue  # a genuinely blank row (trailing empty rows are common) -- not an error
        if name.upper().startswith("EXAMPLE"):
            continue  # the template's own example row, left in by mistake -- silently skip

        barcode = _clean_str(barcode_raw) or None
        unit = _clean_str(unit_raw) or "unit"

        if unit not in _COMMON_UNITS:
            errors.append(
                ImportRowError(
                    row=row_num,
                    field="Unit",
                    message=f'"{unit}" is not one of the allowed units. Use the dropdown.',
                )
            )

        try:
            reorder_point = int(reorder_raw) if reorder_raw is not None else 10
            if reorder_point < 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                ImportRowError(
                    row=row_num, field="Reorder point", message="Must be a whole number, 0 or more."
                )
            )
            reorder_point = 0

        try:
            price = float(price_raw) if price_raw is not None else 0.0
            if price < 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                ImportRowError(
                    row=row_num, field="Selling price", message="Must be a number, 0 or more."
                )
            )
            price = 0.0

        name_key = name.lower()
        if name_key in seen_names:
            errors.append(
                ImportRowError(
                    row=row_num,
                    field="Name",
                    message=f"Duplicate of row {seen_names[name_key]} in this same file.",
                )
            )
        else:
            seen_names[name_key] = row_num

        row_already_invalid = False
        if len(name) > 150:
            errors.append(
                ImportRowError(
                    row=row_num, field="Name", message="Must be 150 characters or fewer."
                )
            )
            row_already_invalid = True
        if barcode and len(barcode) > 64:
            errors.append(
                ImportRowError(
                    row=row_num, field="Barcode", message="Must be 64 characters or fewer."
                )
            )
            row_already_invalid = True

        if barcode:
            if barcode in seen_barcodes:
                errors.append(
                    ImportRowError(
                        row=row_num,
                        field="Barcode",
                        message=f"Duplicate of row {seen_barcodes[barcode]} in this same file.",
                    )
                )
            else:
                seen_barcodes[barcode] = row_num

        # Defensive backstop: even with every check above, construct
        # via try/except rather than trust that this list of checks is
        # exhaustive against every constraint the schema could ever
        # gain. A row that somehow still fails becomes a normal,
        # reported error -- never an unhandled crash.
        if row_already_invalid:
            continue  # already reported above; avoid a duplicate message for the same row

        try:
            candidates.append(
                ProductCreate(
                    name=name,
                    barcode=barcode,
                    unit=unit,
                    reorder_point=reorder_point,
                    default_selling_price=price,
                )
            )
        except ValidationError as exc:
            errors.append(
                ImportRowError(
                    row=row_num, field="Row", message=f"Invalid data: {exc.errors()[0]['msg']}"
                )
            )

    if not candidates:
        errors.append(
            ImportRowError(row=0, field="File", message="No product rows found in this file.")
        )
        return candidates, errors

    # Duplicate check against the EXISTING catalog -- same case-
    # insensitive, active-only rule as single product creation.
    names_to_check = [c.name.lower() for c in candidates]
    existing_names = await db.execute(
        select(func.lower(Product.name)).where(
            func.lower(Product.name).in_(names_to_check), Product.deleted_at.is_(None)
        )
    )
    existing_name_set = {row[0] for row in existing_names.all()}

    barcodes_to_check = [c.barcode for c in candidates if c.barcode]
    existing_barcode_set: set[str] = set()
    if barcodes_to_check:
        existing_barcodes = await db.execute(
            select(Product.barcode).where(
                Product.barcode.in_(barcodes_to_check), Product.deleted_at.is_(None)
            )
        )
        existing_barcode_set = {row[0] for row in existing_barcodes.all()}

    for idx, candidate in enumerate(candidates):
        row_num = idx + 2  # approximate for reporting; exact row tracked above for in-file dupes
        if candidate.name.lower() in existing_name_set:
            errors.append(
                ImportRowError(
                    row=seen_names[candidate.name.lower()],
                    field="Name",
                    message=f'"{candidate.name}" already exists in the catalog.',
                )
            )
        if candidate.barcode and candidate.barcode in existing_barcode_set:
            errors.append(
                ImportRowError(
                    row=seen_barcodes[candidate.barcode],
                    field="Barcode",
                    message=f'Barcode "{candidate.barcode}" already exists in the catalog.',
                )
            )

    return candidates, errors


async def bulk_import(db: AsyncSession, file_bytes: bytes) -> BulkImportResult:
    candidates, errors = await _parse_and_validate(db, file_bytes)

    if errors:
        # All-or-nothing: a rejected file imports exactly zero rows,
        # regardless of how many were individually clean. Reporting
        # every problem at once (not just the first) is what lets one
        # correction pass fix the whole file instead of a slow back-
        # and-forth discovering one bad row per re-upload.
        raise HTTPException(
            status_code=422,
            detail={
                "message": f"{len(errors)} problem(s) found. Nothing was imported.",
                "errors": [e.model_dump() for e in errors],
            },
        )

    for candidate in candidates:
        db.add(Product(**candidate.model_dump()))
    await db.commit()

    return BulkImportResult(created=len(candidates))
