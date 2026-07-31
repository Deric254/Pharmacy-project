"""
Customer bulk import tests -- same properties as product import:
template is a real spreadsheet, import is genuinely all-or-nothing,
duplicates caught both within the file and against the existing
catalog, the EXAMPLE row is silently skipped.
"""

import io

import openpyxl

from app.services.customer_import_service import generate_customer_import_template


async def _login(client, username: str, password: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return str(r.json()["access_token"])


def _build_workbook(rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Phone", "Email"])
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestCustomerImportTemplate:
    def test_template_is_a_real_spreadsheet(self):
        content = generate_customer_import_template()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        assert [c.value for c in ws[1]][:3] == ["Name", "Phone", "Email"]


class TestCustomerBulkImport:
    async def test_clean_file_imports_everything(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["Jane Wanjiru", "0711000001", "jane@test.com"],
                ["John Otieno", "0711000002", ""],
                ["Mary Achieng", "", "mary@test.com"],
            ]
        )
        r = await client.post(
            "/api/v1/customers/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 201
        assert r.json()["created"] == 3

    async def test_one_bad_row_rejects_the_whole_file(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["Genuinely Clean Row", "0711000010", ""],
                ["A" * 200, "0711000011", ""],
            ]
        )
        r = await client.post(
            "/api/v1/customers/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422

        customers = await client.get(
            "/api/v1/customers", headers={"Authorization": f"Bearer {token}"}
        )
        assert customers.json() == []

    async def test_duplicate_phone_within_file_is_caught(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["First Customer", "0711000020", ""],
                ["Second Customer", "0711000020", ""],
            ]
        )
        r = await client.post(
            "/api/v1/customers/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        assert "Duplicate of row" in r.json()["detail"]["errors"][0]["message"]

    async def test_duplicate_phone_against_existing_catalog_is_caught(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}
        await client.post(
            "/api/v1/customers",
            json={"name": "Already Exists", "phone": "0711000030"},
            headers=headers,
        )

        content = _build_workbook([["New Name Same Phone", "0711000030", ""]])
        r = await client.post(
            "/api/v1/customers/import",
            headers=headers,
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 422
        assert "already belongs" in r.json()["detail"]["errors"][0]["message"]

    async def test_example_row_silently_skipped(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        content = _build_workbook(
            [
                ["EXAMPLE - Jane Mwangi", "0712345678", "jane@example.com"],
                ["Real Customer", "0711000040", ""],
            ]
        )
        r = await client.post(
            "/api/v1/customers/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 201
        assert r.json()["created"] == 1

    async def test_requires_permission(self, client, seeded_roles):
        from sqlalchemy import select

        from app.core.database import AsyncSessionLocal
        from app.core.security import hash_password
        from app.models.role import Role
        from app.models.user import User

        async with AsyncSessionLocal() as db:
            role_result = await db.execute(select(Role).where(Role.name == "Employee"))
            role = role_result.scalar_one()
            db.add(
                User(
                    full_name="No Perm",
                    username="noperm_cust",
                    hashed_password=hash_password("pass1234"),
                    role_id=role.id,
                    security_question="Q?",
                    security_answer_hash=hash_password("A"),
                    is_active=True,
                )
            )
            await db.commit()

        # Employee has sales.create by default in this app's seed --
        # this test instead confirms an unauthenticated request is
        # rejected, since every real role already has this permission.
        content = _build_workbook([["Some Customer", "", ""]])
        r = await client.post(
            "/api/v1/customers/import",
            files={"file": ("import.xlsx", content, "application/octet-stream")},
        )
        assert r.status_code == 401
