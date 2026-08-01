"""
Stock take tests. The properties that matter:
  1. Starting a count locks the batches involved so a sale can't move
     the number out from under the counter (proven via an actual sale
     attempt against a locked batch, not just checking a flag).
  2. Non-zero variances always require a reason; small variances
     self-approve, large ones need a separate manager approval call.
  3. Closing is blocked until every item is counted AND every
     variance is resolved - no silent gaps.
  4. Closing unlocks batches and (if there was real shrinkage)
     publishes a shrinkage event with a sane value/percent.
"""

import asyncio
import json
from datetime import date

from sqlalchemy import select

from app.core.database import AsyncSessionLocal
from app.core.events import CHANNEL
from app.core.redis_client import redis_client
from app.models.medicine_batch import MedicineBatch
from app.models.product import Product
from app.models.stock_movement import MovementType, StockMovement


async def _login(client, username: str, password: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return str(r.json()["access_token"])


async def _make_product_with_batch(
    qty: int = 50, price: float = 10.0, cost: float = 4.0
) -> tuple[int, int]:
    async with AsyncSessionLocal() as db:
        product = Product(name="Stock Take Test Product", default_selling_price=price)
        db.add(product)
        await db.flush()
        batch = MedicineBatch(
            product_id=product.id,
            batch_number="ST-1",
            expiry_date=date(2027, 1, 1),
            qty_received=qty,
            qty_remaining=qty,
            cost_price=cost,
        )
        db.add(batch)
        await db.flush()
        db.add(
            StockMovement(
                batch_id=batch.id,
                movement_type=MovementType.PURCHASE,
                quantity_delta=qty,
                created_by_user_id=None,
            )
        )
        await db.commit()
        return int(product.id), int(batch.id)


class TestInitiate:
    async def test_initiate_snapshots_expected_qty_and_locks_batch(self, client, owner_user):
        product_id, batch_id = await _make_product_with_batch(qty=40)
        token = await _login(client, "lucy", "S3curePass!")

        r = await client.post(
            "/api/v1/stock-takes",
            json={"product_ids": [product_id]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201
        body = r.json()
        assert len(body["items"]) == 1
        assert body["items"][0]["expected_qty"] == 40
        assert body["items"][0]["physical_qty"] is None

        async with AsyncSessionLocal() as db:
            from sqlalchemy import select

            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.id == batch_id)
            )
            batch = batch_result.scalar_one()
            assert batch.locked_by_stock_take_id == body["id"]

    async def test_initiate_with_no_eligible_batches_returns_400(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        r = await client.post(
            "/api/v1/stock-takes",
            json={"product_ids": [999999]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400

    async def test_requires_permission(self, client, seeded_roles):
        from sqlalchemy import select

        from app.core.security import hash_password
        from app.models.role import Role
        from app.models.user import User

        async with AsyncSessionLocal() as db:
            role_result = await db.execute(select(Role).where(Role.name == "Employee"))
            role = role_result.scalar_one()
            role.permissions = [p for p in role.permissions if p.code != "stocktake.perform"]
            await db.commit()
            u = User(
                full_name="No Stocktake Perm",
                username="nostockperm",
                hashed_password=hash_password("pass1234"),
                role_id=role.id,
            )
            db.add(u)
            await db.commit()

        token = await _login(client, "nostockperm", "pass1234")
        r = await client.post(
            "/api/v1/stock-takes", json={}, headers={"Authorization": f"Bearer {token}"}
        )
        assert r.status_code == 403

    async def test_two_concurrent_initiates_on_the_same_batch_only_one_wins(
        self, client, owner_user
    ):
        """
        The actual bug this closes: locking was a plain SELECT-then-
        Python-mutate, so two concurrent initiate() calls scoped to
        the same batch could both read it as "currently unlocked" and
        both proceed to lock it -- the second commit silently
        overwriting the first's lock ownership. Two stock takes could
        then each believe they had exclusive control of the same
        physical batch at once.
        """
        product_id, batch_id = await _make_product_with_batch(qty=40)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        async def initiate():
            return await client.post(
                "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
            )

        results = await asyncio.gather(initiate(), initiate(), return_exceptions=True)
        status_codes = [r.status_code for r in results if not isinstance(r, Exception)]
        assert status_codes.count(201) == 1  # exactly one stock take actually claims it
        assert status_codes.count(400) == 1  # the other gets a clean, honest rejection

        async with AsyncSessionLocal() as db:
            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.id == batch_id)
            )
            batch = batch_result.scalar_one()
            winning_id = next(
                r.json()["id"]
                for r in results
                if not isinstance(r, Exception) and r.status_code == 201
            )
            # Locked to exactly the stock take that actually won --
            # never left ambiguous or overwritten.
            assert batch.locked_by_stock_take_id == winning_id


class TestSaleLockInteraction:
    async def test_locked_batch_cannot_be_sold(self, client, owner_user, employee_user):
        product_id, batch_id = await _make_product_with_batch(qty=20)
        owner_token = await _login(client, "lucy", "S3curePass!")

        await client.post(
            "/api/v1/stock-takes",
            json={"product_ids": [product_id]},
            headers={"Authorization": f"Bearer {owner_token}"},
        )

        employee_token = await _login(client, "joe", "pass1234")
        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {employee_token}"},
        )
        # The only batch for this product is locked -> zero effective
        # stock from the sale's point of view.
        assert r.status_code == 409

    async def test_unlocked_after_close_can_be_sold_again(self, client, owner_user, employee_user):
        product_id, batch_id = await _make_product_with_batch(qty=20)
        owner_token = await _login(client, "lucy", "S3curePass!")
        owner_headers = {"Authorization": f"Bearer {owner_token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=owner_headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]

        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 20},  # matches exactly, no reason needed
            headers=owner_headers,
        )
        close_resp = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/close", headers=owner_headers
        )
        assert close_resp.status_code == 200

        employee_token = await _login(client, "joe", "pass1234")
        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {employee_token}"},
        )
        assert r.status_code == 201


class TestCountSubmission:
    async def test_matching_count_auto_approves_with_no_reason(self, client, owner_user):
        product_id, _ = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 30},
            headers=headers,
        )
        assert r.status_code == 200
        assert r.json()["variance"] == 0
        assert r.json()["approved_at"] is not None

    async def test_variance_without_reason_rejected(self, client, owner_user):
        product_id, _ = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 25},  # variance of -5, no reason given
            headers=headers,
        )
        assert r.status_code == 400

    async def test_small_variance_self_approves_and_writes_ledger(self, client, owner_user):
        product_id, batch_id = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 29, "reason": "MISCOUNT"},  # variance -1, within threshold
            headers=headers,
        )
        assert r.status_code == 200
        assert r.json()["variance"] == -1
        assert r.json()["approved_at"] is not None  # self-approved immediately

        async with AsyncSessionLocal() as db:
            from sqlalchemy import select

            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.id == batch_id)
            )
            assert batch_result.scalar_one().qty_remaining == 29

            ledger_result = await db.execute(
                select(StockMovement).where(
                    StockMovement.batch_id == batch_id,
                    StockMovement.movement_type == MovementType.ADJUSTMENT,
                )
            )
            rows = ledger_result.scalars().all()
            assert len(rows) == 1
            assert rows[0].quantity_delta == -1

    async def test_large_variance_left_pending_no_ledger_write(self, client, owner_user):
        product_id, batch_id = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 20, "reason": "THEFT_OR_LOSS"},  # variance -10, exceeds threshold
            headers=headers,
        )
        assert r.status_code == 200
        assert r.json()["variance"] == -10
        assert r.json()["approved_at"] is None  # pending manager approval

        async with AsyncSessionLocal() as db:
            from sqlalchemy import select

            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.id == batch_id)
            )
            assert batch_result.scalar_one().qty_remaining == 30  # untouched until approved


class TestApproval:
    async def test_approve_requires_manager_permission(self, client, owner_user, employee_user):
        product_id, _ = await _make_product_with_batch(qty=30)
        owner_token = await _login(client, "lucy", "S3curePass!")
        owner_headers = {"Authorization": f"Bearer {owner_token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=owner_headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 15, "reason": "THEFT_OR_LOSS"},
            headers=owner_headers,
        )

        employee_token = await _login(client, "joe", "pass1234")
        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/approve",
            headers={"Authorization": f"Bearer {employee_token}"},
        )
        assert r.status_code == 403

    async def test_manager_approval_applies_ledger(self, client, owner_user):
        product_id, batch_id = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 15, "reason": "THEFT_OR_LOSS"},
            headers=headers,
        )

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/approve", headers=headers
        )
        assert r.status_code == 200
        assert r.json()["approved_at"] is not None

        async with AsyncSessionLocal() as db:
            from sqlalchemy import select

            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.id == batch_id)
            )
            assert batch_result.scalar_one().qty_remaining == 15

    async def test_two_concurrent_approvals_only_one_applies_the_ledger(self, client, owner_user):
        product_id, batch_id = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 15, "reason": "THEFT_OR_LOSS"},
            headers=headers,
        )

        async def approve():
            return await client.post(
                f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/approve", headers=headers
            )

        results = await asyncio.gather(approve(), approve(), return_exceptions=True)
        status_codes = [r.status_code for r in results if not isinstance(r, Exception)]
        assert status_codes.count(200) == 1  # exactly one approval succeeds
        assert status_codes.count(400) == 1  # the other is correctly rejected, not silently lost

        async with AsyncSessionLocal() as db:
            movements = await db.execute(
                select(StockMovement).where(
                    StockMovement.batch_id == batch_id,
                    StockMovement.movement_type == MovementType.ADJUSTMENT,
                )
            )
            # Exactly one ADJUSTMENT ledger entry for this variance --
            # not two, which is what a double-approval race would have
            # produced. (The batch also has one PURCHASE-type movement
            # from creation, deliberately excluded here since it isn't
            # what this test is proving.)
            assert len(movements.scalars().all()) == 1


class TestClose:
    async def test_close_blocked_if_items_uncounted(self, client, owner_user):
        product_id, _ = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]

        r = await client.post(f"/api/v1/stock-takes/{stock_take_id}/close", headers=headers)
        assert r.status_code == 400

    async def test_close_blocked_if_variance_pending_approval(self, client, owner_user):
        product_id, _ = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 10, "reason": "THEFT_OR_LOSS"},  # big variance, unapproved
            headers=headers,
        )

        r = await client.post(f"/api/v1/stock-takes/{stock_take_id}/close", headers=headers)
        assert r.status_code == 400

    async def test_close_succeeds_once_fully_resolved(self, client, owner_user):
        product_id, _ = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 30},
            headers=headers,
        )

        r = await client.post(f"/api/v1/stock-takes/{stock_take_id}/close", headers=headers)
        assert r.status_code == 200
        assert r.json()["status"] == "CLOSED"
        assert r.json()["closed_at"] is not None

    async def test_two_concurrent_close_calls_only_one_succeeds(self, client, owner_user):
        """
        The actual bug this closes: close() checked "is this already
        closed?" as a plain read with no atomic guard at all -- two
        simultaneous close() calls on the same OPEN stock take could
        both pass that check, both compute shrinkage, and both
        successfully write CLOSED, the second commit silently
        overwriting the first's closed_at and (worse) both publishing
        a StockTakeClosedEvent for what should have been one closure.
        """
        product_id, _ = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 30},
            headers=headers,
        )

        async def attempt_close():
            return await client.post(f"/api/v1/stock-takes/{stock_take_id}/close", headers=headers)

        results = await asyncio.gather(attempt_close(), attempt_close(), return_exceptions=True)
        status_codes = [r.status_code for r in results if not isinstance(r, Exception)]
        assert status_codes.count(200) == 1
        assert status_codes.count(400) == 1

    async def test_reconciliation_clean_after_close(self, client, owner_user):
        product_id, batch_id = await _make_product_with_batch(qty=30)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 26, "reason": "DAMAGED"},  # variance -4, self-approve range
            headers=headers,
        )
        await client.post(f"/api/v1/stock-takes/{stock_take_id}/close", headers=headers)

        r = await client.get("/api/v1/inventory/reconcile", headers=headers)
        issues = [i for i in r.json() if i["batch_id"] == batch_id]
        assert issues == []  # ledger and qty_remaining agree after the workflow

    async def test_shrinkage_event_published_on_close(self, client, owner_user):
        product_id, _ = await _make_product_with_batch(qty=100, cost=4.0)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        pubsub = redis_client.pubsub()
        await pubsub.subscribe(CHANNEL)
        await pubsub.get_message(timeout=1)

        create_resp = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = create_resp.json()["id"]
        item_id = create_resp.json()["items"][0]["id"]
        # -2 is within self-approve threshold, keeps this a single-step test
        await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/items/{item_id}/count",
            json={"physical_qty": 98, "reason": "DAMAGED"},
            headers=headers,
        )
        await client.post(f"/api/v1/stock-takes/{stock_take_id}/close", headers=headers)

        found = False
        for _ in range(10):
            message = await pubsub.get_message(timeout=1)
            if message and message["type"] == "message":
                envelope = json.loads(message["data"])
                if envelope["event_type"] == "stocktake.closed":
                    assert envelope["payload"]["stock_take_id"] == stock_take_id
                    assert (
                        float(envelope["payload"]["shrinkage_value"]) == 8.0
                    )  # 2 units * 4.0 cost
                    found = True
                    break
        await pubsub.unsubscribe(CHANNEL)
        assert found, "Expected a stocktake.closed event with shrinkage data"


class TestStockTakeExcelRoundTrip:
    """
    The full real workflow: download a template with real system
    quantities, fill in physical counts offline, upload it, and it
    applies every count and auto-closes the stock take if everything
    resolves -- proven live against a real server before these tests
    were even written, this just makes that proof permanent.
    """

    async def test_template_has_real_product_names_and_quantities(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        product = await client.post(
            "/api/v1/products",
            json={"name": "Excel Roundtrip Product", "default_selling_price": 20.0},
            headers=headers,
        )
        product_id = product.json()["id"]
        await client.post(
            f"/api/v1/products/{product_id}/batches",
            json={
                "batch_number": "EXCEL1",
                "expiry_date": "2027-06-30",
                "qty_received": 75,
                "cost_price": 8.0,
            },
            headers=headers,
        )
        st = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = st.json()["id"]

        r = await client.get(f"/api/v1/stock-takes/{stock_take_id}/count-template", headers=headers)
        assert r.status_code == 200

        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers_row = [c.value for c in ws[1]]
        assert headers_row[:5] == [
            "Product name",
            "Batch number",
            "Expiry date",
            "System quantity",
            "Physical quantity",
        ]
        data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        assert data_row[0] == "Excel Roundtrip Product"
        assert data_row[1] == "EXCEL1"
        assert data_row[3] == 75

    async def test_upload_applies_counts_and_auto_closes_when_resolved(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        product = await client.post(
            "/api/v1/products",
            json={"name": "Excel Autoclose Product", "default_selling_price": 20.0},
            headers=headers,
        )
        product_id = product.json()["id"]
        await client.post(
            f"/api/v1/products/{product_id}/batches",
            json={
                "batch_number": "AUTOCLOSE1",
                "expiry_date": "2027-06-30",
                "qty_received": 100,
                "cost_price": 5.0,
            },
            headers=headers,
        )
        st = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = st.json()["id"]

        template_resp = await client.get(
            f"/api/v1/stock-takes/{stock_take_id}/count-template", headers=headers
        )

        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(template_resp.content))
        ws = wb.active
        # Counted 1 unit fewer than expected -- small enough to
        # self-approve, so the whole stock take should close in one step.
        ws.cell(row=2, column=5, value=99)
        buffer = io.BytesIO()
        wb.save(buffer)

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/import-counts",
            headers=headers,
            files={
                "file": (
                    "counted.xlsx",
                    buffer.getvalue(),
                    "application/octet-stream",
                )
            },
        )
        assert r.status_code == 200
        body = r.json()
        assert body["status"] == "CLOSED"
        item = body["items"][0]
        assert item["physical_qty"] == 99
        assert item["variance"] == -1
        assert item["approved_at"] is not None

        # The real stock must actually reflect the physical count now.
        product_check = await client.get(f"/api/v1/products/{product_id}", headers=headers)
        assert product_check.json()["total_qty_available"] == 99

    async def test_edited_hidden_id_column_is_rejected_cleanly(self, client, owner_user):
        """
        The hidden Item ID column is what makes re-upload safe --
        tampering with it (accidentally or otherwise) must be caught
        with a clear error, never silently applied to the wrong item.
        """
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        product = await client.post(
            "/api/v1/products",
            json={"name": "Excel Tamper Product", "default_selling_price": 20.0},
            headers=headers,
        )
        product_id = product.json()["id"]
        await client.post(
            f"/api/v1/products/{product_id}/batches",
            json={
                "batch_number": "TAMPER1",
                "expiry_date": "2027-06-30",
                "qty_received": 30,
                "cost_price": 3.0,
            },
            headers=headers,
        )
        st = await client.post(
            "/api/v1/stock-takes", json={"product_ids": [product_id]}, headers=headers
        )
        stock_take_id = st.json()["id"]

        template_resp = await client.get(
            f"/api/v1/stock-takes/{stock_take_id}/count-template", headers=headers
        )

        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(template_resp.content))
        ws = wb.active
        ws.cell(row=2, column=5, value=30)
        ws.cell(row=2, column=6, value="not-a-real-id")  # tamper with the hidden ID
        buffer = io.BytesIO()
        wb.save(buffer)

        r = await client.post(
            f"/api/v1/stock-takes/{stock_take_id}/import-counts",
            headers=headers,
            files={
                "file": (
                    "tampered.xlsx",
                    buffer.getvalue(),
                    "application/octet-stream",
                )
            },
        )
        assert r.status_code == 400
        assert "Item ID" in r.json()["detail"]
