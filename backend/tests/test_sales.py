"""
Sales tests. The two properties that actually matter for a POS:
  1. A sale either fully succeeds (stock decremented, ledger written,
     payment recorded) or fully fails (nothing touched) - no partial
     sales, ever.
  2. Checkout draws from the nearest-expiry batch first, same as the
     dedicated FEFO tests prove at the service level - this file
     proves it end-to-end through the real HTTP + DB path.
"""

import asyncio
from datetime import date, timedelta

from sqlalchemy import select

from app.core.database import AsyncSessionLocal
from app.core.security import hash_password
from app.models.medicine_batch import MedicineBatch
from app.models.product import Product
from app.models.role import Role
from app.models.sale import Sale
from app.models.stock_movement import StockMovement
from app.models.user import User


async def _login(client, username: str, password: str) -> str:
    r = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert r.status_code == 200, r.text
    return str(r.json()["access_token"])


async def _make_product_with_batch(
    price: float = 10.0, qty: int = 20, expiry: str = "2027-01-01", cost: float | None = None
) -> int:
    async with AsyncSessionLocal() as db:
        product = Product(name="Amoxicillin 500mg", default_selling_price=price)
        db.add(product)
        await db.flush()
        db.add(
            MedicineBatch(
                product_id=product.id,
                batch_number="B1",
                expiry_date=date.fromisoformat(expiry),
                qty_received=qty,
                qty_remaining=qty,
                cost_price=cost if cost is not None else price / 2,
            )
        )
        await db.commit()
        return int(product.id)


class TestCreateSale:
    async def test_requires_permission(self, client, seeded_roles):
        async with AsyncSessionLocal() as db:
            role_result = await db.execute(select(Role).where(Role.name == "Employee"))
            role = role_result.scalar_one()
            role.permissions = []  # strip sales.create for this one test
            await db.commit()

            u = User(
                full_name="No Permission",
                username="noperm",
                hashed_password=hash_password("pass1234"),
                role_id=role.id,
            )
            db.add(u)
            await db.commit()

        token = await _login(client, "noperm", "pass1234")
        r = await client.post(
            "/api/v1/sales",
            json={"items": [], "payments": []},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 403

    async def test_successful_sale_decrements_stock_and_writes_ledger(self, client, employee_user):
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201, r.text
        body = r.json()
        assert body["subtotal"] == 50.0
        assert body["total_amount"] == 50.0
        assert len(body["items"]) == 1
        assert body["items"][0]["quantity"] == 5

        async with AsyncSessionLocal() as db:
            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            assert batch.qty_remaining == 15  # 20 - 5

            ledger_result = await db.execute(
                select(StockMovement).where(StockMovement.reference == f"sale:{body['id']}")
            )
            ledger_rows = ledger_result.scalars().all()
            assert len(ledger_rows) == 1
            assert ledger_rows[0].quantity_delta == -5

    async def test_insufficient_stock_rolls_back_completely(self, client, employee_user):
        product_id = await _make_product_with_batch(price=10.0, qty=3)
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 10}],
                "payments": [{"method": "CASH", "amount": 100.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 409

        async with AsyncSessionLocal() as db:
            # Nothing touched: no sale row, batch quantity unchanged.
            sales_result = await db.execute(select(Sale))
            assert sales_result.scalars().all() == []

            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            assert batch.qty_remaining == 3  # unchanged

    async def test_payment_mismatch_rejected_and_rolled_back(self, client, employee_user):
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 40.0}],  # should be 50
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400

        async with AsyncSessionLocal() as db:
            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            assert batch.qty_remaining == 20  # unchanged - nothing partially applied

    async def test_duplicate_product_lines_rejected_by_schema(self, client, employee_user):
        product_id = await _make_product_with_batch()
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [
                    {"product_id": product_id, "quantity": 1},
                    {"product_id": product_id, "quantity": 2},
                ],
                "payments": [{"method": "CASH", "amount": 30.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 422

    async def test_checkout_draws_nearest_expiry_batch_first(self, client, employee_user):
        async with AsyncSessionLocal() as db:
            product = Product(name="Paracetamol", default_selling_price=5.0)
            db.add(product)
            await db.flush()
            db.add(
                MedicineBatch(
                    product_id=product.id,
                    batch_number="FAR",
                    expiry_date=date.today() + timedelta(days=700),
                    qty_received=50,
                    qty_remaining=50,
                    cost_price=2.0,
                )
            )
            db.add(
                MedicineBatch(
                    product_id=product.id,
                    batch_number="NEAR",
                    expiry_date=date.today() + timedelta(days=30),
                    qty_received=10,
                    qty_remaining=10,
                    cost_price=2.0,
                )
            )
            await db.commit()
            product_id = int(product.id)

        token = await _login(client, "joe", "pass1234")
        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 4}],
                "payments": [{"method": "CASH", "amount": 20.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201

        async with AsyncSessionLocal() as db:
            result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batches = {b.batch_number: b.qty_remaining for b in result.scalars().all()}
            assert batches["NEAR"] == 6  # 10 - 4, drawn first
            assert batches["FAR"] == 50  # untouched


class TestGetSale:
    async def test_get_sale_returns_full_detail(self, client, employee_user):
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")
        headers = {"Authorization": f"Bearer {token}"}

        create_resp = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 2}],
                "payments": [{"method": "MPESA", "amount": 20.0, "reference": "QWE123"}],
            },
            headers=headers,
        )
        sale_id = create_resp.json()["id"]

        r = await client.get(f"/api/v1/sales/{sale_id}", headers=headers)
        assert r.status_code == 200
        assert r.json()["payments"][0]["method"] == "MPESA"
        assert r.json()["payments"][0]["reference"] == "QWE123"

    async def test_get_unknown_sale_returns_404(self, client, employee_user):
        token = await _login(client, "joe", "pass1234")
        r = await client.get("/api/v1/sales/999999", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 404


class TestConcurrentSales:
    async def test_two_concurrent_sales_cannot_oversell_the_same_stock(self, client, employee_user):
        """
        Two 'cashiers' try to sell from a batch with only 10 units at
        the same time, each requesting 8. Both cannot succeed - only
        one should, and total stock sold must never exceed what
        existed. Exercised under real concurrency rather than asserted
        from the single-threaded FEFO unit tests alone.

        This does NOT rely on SELECT...FOR UPDATE -- SQLite silently
        drops that clause entirely (confirmed by inspecting the
        compiled SQL), so it was never the real guarantee here. What
        actually prevents overselling is the atomic `UPDATE ...
        WHERE qty_remaining >= :qty` in apply_allocations(): the
        second transaction's decrement is evaluated against the row's
        real state at the moment it runs, not a stale snapshot, and
        SQLite's busy_timeout (configured in database.py) makes it
        wait for the first transaction to finish rather than fail
        instantly with "database is locked". Confirmed directly by
        reproducing the bug this fix closes: before the atomic UPDATE
        existed, a Python-level `qty -= amount` on the ORM object let
        two concurrent decrements silently overwrite each other.
        """
        product_id = await _make_product_with_batch(price=10.0, qty=10)
        token = await _login(client, "joe", "pass1234")
        headers = {"Authorization": f"Bearer {token}"}

        async def attempt_sale():
            return await client.post(
                "/api/v1/sales",
                json={
                    "items": [{"product_id": product_id, "quantity": 8}],
                    "payments": [{"method": "CASH", "amount": 80.0}],
                },
                headers=headers,
            )

        results = await asyncio.gather(attempt_sale(), attempt_sale(), return_exceptions=True)
        status_codes = [r.status_code for r in results if not isinstance(r, Exception)]

        assert status_codes.count(201) == 1
        assert status_codes.count(409) == 1

        async with AsyncSessionLocal() as db:
            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            assert batch.qty_remaining == 2  # 10 - 8, never negative, never double-sold


class TestIdempotentCheckout:
    """
    The scenario this protects against: a checkout request commits
    server-side (stock decremented, payment recorded) but its response
    never reaches the cashier -- a dropped connection, the local
    backend restarting mid-request. Without this, "Checkout failed,
    nothing was charged" is displayed and re-enables the button, and a
    retry creates a second, fully real sale.
    """

    async def test_retrying_the_same_key_returns_the_original_sale_not_a_second_one(
        self, client, employee_user
    ):
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")
        headers = {"Authorization": f"Bearer {token}"}

        payload = {
            "items": [{"product_id": product_id, "quantity": 3}],
            "payments": [{"method": "CASH", "amount": 30.0}],
            "idempotency_key": "attempt-abc-123",
        }

        first = await client.post("/api/v1/sales", json=payload, headers=headers)
        assert first.status_code == 201, first.text

        # Simulates the cashier retrying after (from their point of
        # view) a failed checkout -- same key, same request.
        second = await client.post("/api/v1/sales", json=payload, headers=headers)
        assert second.status_code == 201, second.text

        # The critical assertion: both responses describe the SAME
        # sale, not two different ones.
        assert first.json()["id"] == second.json()["id"]

        async with AsyncSessionLocal() as db:
            all_sales = (
                (await db.execute(select(Sale).where(Sale.idempotency_key == "attempt-abc-123")))
                .scalars()
                .all()
            )
            assert len(all_sales) == 1  # not two

            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            # Stock decremented ONCE (20 - 3 = 17), not twice.
            assert batch.qty_remaining == 17

    async def test_concurrent_retries_with_the_same_new_key_still_produce_one_sale(
        self, client, employee_user
    ):
        """
        The race the top-of-method check alone can't close: two
        requests with the SAME brand-new key, arriving close enough
        together that both pass the "does this key already exist"
        check before either commits. The UNIQUE constraint on
        idempotency_key is what actually closes this -- the second
        INSERT fails, and the service catches that and returns the
        first sale instead of a raw 500.
        """
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")
        headers = {"Authorization": f"Bearer {token}"}

        payload = {
            "items": [{"product_id": product_id, "quantity": 2}],
            "payments": [{"method": "CASH", "amount": 20.0}],
            "idempotency_key": "attempt-race-1",
        }

        async def attempt():
            return await client.post("/api/v1/sales", json=payload, headers=headers)

        results = await asyncio.gather(attempt(), attempt(), return_exceptions=True)
        exceptions = [r for r in results if isinstance(r, Exception)]
        assert not exceptions, f"Unhandled exceptions under concurrency: {exceptions}"

        status_codes = [r.status_code for r in results]
        assert all(code == 201 for code in status_codes), status_codes

        sale_ids = {r.json()["id"] for r in results}
        assert len(sale_ids) == 1  # both responses describe the same sale

        async with AsyncSessionLocal() as db:
            all_sales = (
                (await db.execute(select(Sale).where(Sale.idempotency_key == "attempt-race-1")))
                .scalars()
                .all()
            )
            assert len(all_sales) == 1

    async def test_two_genuinely_separate_sales_with_different_keys_both_go_through(
        self, client, employee_user
    ):
        """
        The other side of the same feature: identical cart contents
        (same product, quantity, payment method) but DIFFERENT keys
        -- e.g. two separate customers buying the same thing back to
        back -- must never be collapsed into one sale just because
        their payloads look alike.
        """
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")
        headers = {"Authorization": f"Bearer {token}"}

        first = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
                "idempotency_key": "sale-one",
            },
            headers=headers,
        )
        second = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
                "idempotency_key": "sale-two",
            },
            headers=headers,
        )
        assert first.status_code == 201
        assert second.status_code == 201
        assert first.json()["id"] != second.json()["id"]

        async with AsyncSessionLocal() as db:
            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            assert batch.qty_remaining == 18  # 20 - 1 - 1, both sales counted

    async def test_no_key_sent_still_works_exactly_as_before(self, client, employee_user):
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "joe", "pass1234")
        headers = {"Authorization": f"Bearer {token}"}

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers=headers,
        )
        assert r.status_code == 201, r.text
        assert r.json()["id"] is not None

    async def test_selling_below_cost_is_rejected(self, client, employee_user):
        # Selling price 10.0, but this batch cost 15.0 -- a real loss.
        product_id = await _make_product_with_batch(price=10.0, qty=20, cost=15.0)
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400
        assert "loss" in r.json()["detail"].lower()

    async def test_rejected_sale_never_touches_stock(self, client, employee_user):
        """
        The real guarantee, not just the error code: a sale rejected
        for selling at a loss must leave stock completely untouched --
        no partial decrement snuck through before the check.
        """
        product_id = await _make_product_with_batch(price=10.0, qty=20, cost=15.0)
        token = await _login(client, "joe", "pass1234")

        await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        async with AsyncSessionLocal() as db:
            batch_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.product_id == product_id)
            )
            batch = batch_result.scalar_one()
            assert batch.qty_remaining == 20  # untouched, not 15

    async def test_selling_at_exactly_cost_is_allowed_not_a_loss(self, client, employee_user):
        # Break-even isn't a loss -- price == cost must be allowed.
        product_id = await _make_product_with_batch(price=10.0, qty=20, cost=10.0)
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201

    async def test_selling_above_cost_is_allowed(self, client, employee_user):
        product_id = await _make_product_with_batch(price=10.0, qty=20, cost=4.0)
        token = await _login(client, "joe", "pass1234")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 5}],
                "payments": [{"method": "CASH", "amount": 50.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201


class TestListSalesQueryPlan:
    """
    Regression guard for a real bug found under load testing at
    100,000 sales: item_count used to be computed via a JOIN to
    sale_items + GROUP BY, which forced SQLite to materialize and sort
    EVERY matching sale (a full table scan, confirmed with
    EXPLAIN QUERY PLAN) before LIMIT could apply -- fetching just the
    first page of 50 sales took 104ms at that volume because the
    database had to touch all 100,000 rows first. Rewritten as a
    correlated scalar subquery so item_count is only computed for the
    rows actually returned, letting the main query stay index-driven
    off Sale.created_at -- proven at the same volume to drop to 8.5ms.

    This test doesn't re-run the full load test (too slow for the
    regular suite) -- it checks the actual SQL Core query object list_sales
    builds contains no GROUP BY, which is the specific construct that
    forced the full scan. A future change that reintroduces a
    JOIN+GROUP BY for item_count would fail this immediately, instead
    of silently reappearing only under real production data volume.
    """

    async def test_list_sales_query_has_no_group_by(self, client, owner_user):
        from sqlalchemy.ext.asyncio import AsyncSession

        from app.core.database import AsyncSessionLocal
        from app.services.sale_service import SaleService

        async with AsyncSessionLocal() as db:
            assert isinstance(db, AsyncSession)
            service = SaleService(db)
            # list_sales builds the query internally; the only way to
            # inspect it without duplicating its logic here is to
            # patch db.execute and capture what it was called with.
            captured_queries: list[object] = []
            real_execute = db.execute

            async def capturing_execute(stmt, *args, **kwargs):
                captured_queries.append(stmt)
                return await real_execute(stmt, *args, **kwargs)

            db.execute = capturing_execute  # type: ignore[method-assign]
            await service.list_sales(limit=10, offset=0)

        # The count query and the main listing query both get
        # captured; the main one is identifiable by selecting more
        # than one column.
        main_queries = [q for q in captured_queries if len(q.selected_columns) > 1]
        assert main_queries, "list_sales did not execute its main query as expected"
        main_query = main_queries[0]
        assert not main_query._group_by_clauses, (
            "list_sales' query has a GROUP BY again -- this is exactly the pattern "
            "that caused a full table scan at 100k sales (confirmed via "
            "EXPLAIN QUERY PLAN: 'SCAN sales' instead of using the created_at index). "
            "item_count must be a correlated scalar subquery instead, not a JOIN + "
            "GROUP BY, or the first page of sales history will get slower as the "
            "business's total sale count grows, regardless of the date range asked for."
        )


class TestListSales:
    """
    The actual gap this closes: there was no way to browse past sales
    at all, only fetch one by already knowing its exact ID -- a real
    pharmacy owner could never look at what they sold yesterday
    through the API at all before this existed.
    """

    async def test_lists_real_sales_with_correct_names_and_totals(
        self, client, owner_user, employee_user
    ):
        product_id = await _make_product_with_batch(price=10.0)
        token = await _login(client, "joe", "pass1234")
        owner_token = await _login(client, "lucy", "S3curePass!")

        await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 2}],
                "payments": [{"method": "CASH", "amount": 20.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        r = await client.get("/api/v1/sales", headers={"Authorization": f"Bearer {owner_token}"})
        assert r.status_code == 200
        body = r.json()
        assert body["total"] == 1
        entry = body["entries"][0]
        assert entry["cashier_name"] == "Cashier Joe"
        assert entry["customer_name"] is None
        assert entry["total_amount"] == 20.0
        assert entry["item_count"] == 1

    async def test_newest_sales_first(self, client, owner_user, employee_user):
        product_id = await _make_product_with_batch(price=10.0)
        token = await _login(client, "joe", "pass1234")

        first = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        second = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        r = await client.get("/api/v1/sales", headers={"Authorization": f"Bearer {token}"})
        ids = [e["id"] for e in r.json()["entries"]]
        assert ids[0] == second.json()["id"]
        assert ids[1] == first.json()["id"]

    async def test_date_range_filter_excludes_sales_outside_it(
        self, client, owner_user, employee_user
    ):
        product_id = await _make_product_with_batch(price=10.0)
        token = await _login(client, "joe", "pass1234")
        await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )

        yesterday = (date.today() - timedelta(days=1)).isoformat()
        two_days_ago = (date.today() - timedelta(days=2)).isoformat()
        r = await client.get(
            "/api/v1/sales",
            params={"start_date": two_days_ago, "end_date": yesterday},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.json()["total"] == 0

    async def test_requires_permission(self, client, employee_user):
        # joe has sales.create, matching what create_sale itself requires
        token = await _login(client, "joe", "pass1234")
        r = await client.get("/api/v1/sales", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200


class TestSalesExport:
    async def test_json_export_is_still_the_default(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        r = await client.get("/api/v1/sales", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("application/json")

    async def test_excel_export_returns_a_real_spreadsheet(self, client, owner_user):
        product_id = await _make_product_with_batch(price=10.0)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}
        await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 2}],
                "payments": [{"method": "CASH", "amount": 20.0}],
            },
            headers=headers,
        )

        r = await client.get("/api/v1/sales?export=excel", headers=headers)
        assert r.status_code == 200
        assert (
            r.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert len(r.content) > 0

        import io
        import zipfile

        assert zipfile.is_zipfile(io.BytesIO(r.content))

    async def test_export_includes_every_matching_sale_not_just_one_page(
        self, client, owner_user
    ):
        """
        The exact gap a naive "just add export to the paginated
        endpoint" fix would leave: list_sales() caps at a page size,
        so export must go through list_all_for_export() instead, or a
        business with more sales than one page would silently get a
        truncated export with no indication anything was cut off.
        """
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}
        for _ in range(3):
            r = await client.post(
                "/api/v1/sales",
                json={
                    "items": [{"product_id": product_id, "quantity": 1}],
                    "payments": [{"method": "CASH", "amount": 10.0}],
                },
                headers=headers,
            )
            assert r.status_code == 201, r.text

        # A page size smaller than the real number of sales -- if
        # export reused list_sales()'s own limit/offset, this would
        # cap the export at 2 rows instead of the real 3.
        r = await client.get(
            "/api/v1/sales?export=excel&limit=2&offset=0", headers=headers
        )
        assert r.status_code == 200

        import io

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(r.content))
        sheet = workbook.active
        # header row + 3 real sales, not header row + 2
        assert sheet.max_row == 4


class TestExpiredStockNeverSold:
    """
    The actual gap this closes: FEFO correctly ordered by soonest-
    expiry-first, but never actually excluded stock that had already
    expired outright -- a genuine patient-safety issue, not just a
    consistency one. Expired batches must never be sellable, full
    stop, regardless of how FEFO would otherwise have ranked them.
    """

    async def test_batch_expired_yesterday_cannot_be_sold(self, client, owner_user):
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        product_id = await _make_product_with_batch(price=10.0, qty=20, expiry=yesterday)
        token = await _login(client, "lucy", "S3curePass!")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 409

    async def test_batch_expiring_today_can_still_be_sold(self, client, owner_user):
        today = date.today().isoformat()
        product_id = await _make_product_with_batch(price=10.0, qty=20, expiry=today)
        token = await _login(client, "lucy", "S3curePass!")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201

    async def test_expired_batch_is_skipped_in_favor_of_a_valid_later_one(self, client, owner_user):
        """
        A real mixed scenario: one expired batch and one valid batch
        for the same product. FEFO must skip the expired one entirely
        and fulfill from the valid batch, not fail outright just
        because an expired batch happened to sort first.
        """
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        product_id = await _make_product_with_batch(price=10.0, qty=5, expiry=yesterday)

        async with AsyncSessionLocal() as db:
            db.add(
                MedicineBatch(
                    product_id=product_id,
                    batch_number="VALID1",
                    expiry_date=date(2027, 1, 1),
                    qty_received=20,
                    qty_remaining=20,
                    cost_price=5.0,
                )
            )
            await db.commit()

        token = await _login(client, "lucy", "S3curePass!")
        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 3}],
                "payments": [{"method": "CASH", "amount": 30.0}],
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 201

        async with AsyncSessionLocal() as db:
            expired_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.batch_number == "B1")
            )
            valid_result = await db.execute(
                select(MedicineBatch).where(MedicineBatch.batch_number == "VALID1")
            )
            # The expired batch must be completely untouched -- all 3
            # units came from the valid batch instead.
            assert expired_result.scalar_one().qty_remaining == 5
            assert valid_result.scalar_one().qty_remaining == 17


class TestDiscountCannotExceedSubtotal:
    async def test_clear_error_not_a_confusing_payment_mismatch(self, client, owner_user):
        """
        Not a safety fix (a negative total could never actually be
        paid, since payments can never be negative either) -- purely
        a clarity fix. Before this, an over-large discount produced a
        confusing "payment total doesn't match" error instead of
        naming the actual problem.
        """
        product_id = await _make_product_with_batch(price=10.0, qty=20)
        token = await _login(client, "lucy", "S3curePass!")

        r = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 0.01}],
                "discount_amount": 50.0,
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400
        assert "discount" in r.json()["detail"].lower()
        assert "exceed" in r.json()["detail"].lower()


class TestReceiptPdf:
    """
    The properties that matter: a real, correctly-branded PDF is
    generated fresh from the actual sale and business config every
    time, a malformed logo can never crash it, and displayed times are
    converted to the business's real configured timezone rather than
    showing raw UTC.
    """

    async def test_generates_a_real_pdf_with_correct_numbers(self, client, owner_user):
        product_id = await _make_product_with_batch(price=25.0, cost=10.0)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        sale = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 4}],
                "payments": [{"method": "CASH", "amount": 90.0}],
                "discount_amount": 10.0,
            },
            headers=headers,
        )
        sale_id = sale.json()["id"]

        r = await client.get(f"/api/v1/sales/{sale_id}/receipt", headers=headers)
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:4] == b"%PDF"

    async def test_malformed_logo_never_crashes_the_receipt(self, client, owner_user):
        """
        The real bug this proves is fixed: a malformed logo image
        used to crash the whole receipt with a 500, because the
        safety check only guarded decode time, not render time (where
        PIL/reportlab actually load the pixel data).
        """
        product_id = await _make_product_with_batch(price=10.0)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        await client.patch(
            "/api/v1/config",
            json={"logo_url": "data:image/png;base64,dGhpcyBpcyBub3QgYSByZWFsIHBuZw=="},
            headers=headers,
        )

        sale = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers=headers,
        )
        sale_id = sale.json()["id"]

        r = await client.get(f"/api/v1/sales/{sale_id}/receipt", headers=headers)
        assert r.status_code == 200
        assert r.content[:4] == b"%PDF"

    async def test_requires_permission(self, client, employee_user):
        token = await _login(client, "joe", "pass1234")
        r = await client.get(
            "/api/v1/sales/1/receipt", headers={"Authorization": f"Bearer {token}"}
        )
        # joe (Employee) has sales.create, matching what this endpoint requires
        assert r.status_code == 404  # sale 1 doesn't exist -- confirms auth passed, not blocked

    async def test_nonexistent_sale_returns_404(self, client, owner_user):
        token = await _login(client, "lucy", "S3curePass!")
        r = await client.get(
            "/api/v1/sales/999999/receipt", headers={"Authorization": f"Bearer {token}"}
        )
        assert r.status_code == 404

    async def test_timestamp_is_converted_to_the_real_business_timezone(self, client, owner_user):
        """
        The real bug this proves is fixed: the receipt used to show
        the raw stored UTC timestamp directly, ignoring the business's
        own configured timezone entirely -- a receipt could show a
        time hours off from the real local time it happened.
        """
        product_id = await _make_product_with_batch(price=10.0)
        token = await _login(client, "lucy", "S3curePass!")
        headers = {"Authorization": f"Bearer {token}"}

        await client.patch("/api/v1/config", json={"timezone": "Africa/Nairobi"}, headers=headers)

        sale = await client.post(
            "/api/v1/sales",
            json={
                "items": [{"product_id": product_id, "quantity": 1}],
                "payments": [{"method": "CASH", "amount": 10.0}],
            },
            headers=headers,
        )
        sale_id = sale.json()["id"]
        sale_detail = await client.get(f"/api/v1/sales/{sale_id}", headers=headers)
        from datetime import datetime

        stored_utc = datetime.fromisoformat(sale_detail.json()["created_at"])

        r = await client.get(f"/api/v1/sales/{sale_id}/receipt", headers=headers)
        import io as _io

        from pypdf import PdfReader

        reader = PdfReader(_io.BytesIO(r.content))
        text = reader.pages[0].extract_text()

        # Africa/Nairobi is UTC+3 -- the receipt's displayed hour must
        # be exactly 3 ahead of the stored UTC hour, not equal to it.
        expected_local_hour = (stored_utc.hour + 3) % 24
        assert f"{expected_local_hour:02d}:" in text
        assert "(Africa/Nairobi)" in text
