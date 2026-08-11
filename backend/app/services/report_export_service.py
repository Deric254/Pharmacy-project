"""
Report export utilities.

Both functions take the same shape of input (headers + rows of plain
values) so any report's data can be exported without each report
endpoint reimplementing file generation. Kept intentionally simple --
one sheet, one table -- rather than a templating system, since that's
what an SME pharmacy report actually needs.
"""

import io
from typing import TYPE_CHECKING, Literal

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

if TYPE_CHECKING:
    from reportlab.graphics.shapes import Drawing

    from app.schemas.reports import RevenueTrendPoint, TopCustomerEntry, TopProductEntry

ExportFormat = Literal["json", "excel", "pdf"]

_EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MEDIA_TYPE = "application/pdf"


def export_to_excel(
    headers: list[str], rows: list[list[object]], sheet_title: str = "Report"
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None  # a freshly created Workbook always has an active sheet
    sheet.title = sheet_title[:31]  # Excel sheet name length limit

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)

    column_count = len(headers)
    for col_index in range(1, column_count + 1):
        column_letter = get_column_letter(col_index)
        header_length = len(str(headers[col_index - 1]))
        data_lengths = (
            len(str(row[col_index - 1])) for row in rows if row[col_index - 1] is not None
        )
        max_length = max([header_length, *data_lengths], default=10)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_to_pdf(title: str, headers: list[str], rows: list[list[object]]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title=title, topMargin=3 * cm)

    table_data = [headers] + [[str(cell) for cell in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    def draw_header(canvas: object, _doc: object) -> None:
        _draw_title(canvas, title)

    doc.build([table], onFirstPage=draw_header, onLaterPages=lambda c, d: None)
    return buffer.getvalue()


def _draw_title(canvas: object, title: str) -> None:
    canvas.saveState()  # type: ignore[attr-defined]
    canvas.setFont("Helvetica-Bold", 14)  # type: ignore[attr-defined]
    canvas.drawString(2 * cm, 28 * cm, title)  # type: ignore[attr-defined]
    canvas.restoreState()  # type: ignore[attr-defined]


def build_export_response(
    export: ExportFormat,
    json_payload: object,
    title: str,
    headers: list[str],
    rows: list[list[object]],
) -> object:
    """
    Shared across every export-capable endpoint (reports, and any raw
    data list -- Products, Customers, Audit Trail), so each one just
    supplies its own headers/rows rather than reimplementing this
    branch. Excel filename is derived from the title, sanitized to
    ASCII alphanumerics/spaces/hyphens -- title strings come from
    business-facing labels a user chose, not developer-controlled
    constants, so this can't be trusted blindly as a filesystem path
    component.
    """
    safe_title = "".join(c for c in title if c.isalnum() or c in " -_")[:100] or "Export"
    if export == "excel":
        content = export_to_excel(headers, rows, sheet_title=title)
        return Response(
            content=content,
            media_type=_EXCEL_MEDIA_TYPE,
            headers={"Content-Disposition": f'attachment; filename="{safe_title}.xlsx"'},
        )
    if export == "pdf":
        content = export_to_pdf(title, headers, rows)
        return Response(
            content=content,
            media_type=_PDF_MEDIA_TYPE,
            headers={"Content-Disposition": f'attachment; filename="{safe_title}.pdf"'},
        )
    return json_payload


def generate_profit_loss_pdf(
    business_name: str,
    start_date: str,
    end_date: str,
    revenue: float,
    cost_of_goods_sold: float,
    gross_profit: float,
    gross_margin_percent: float,
    currency: str,
    trend_points: list["RevenueTrendPoint"] | None = None,
    top_products: list["TopProductEntry"] | None = None,
    top_customers: list["TopCustomerEntry"] | None = None,
) -> bytes:
    """
    A real, honest Gross Profit statement -- not a full P&L. This
    system has no expense-tracking module anywhere (no rent, salaries,
    utilities, or any other overhead is recorded), so a true net-profit
    P&L cannot be honestly produced. Rather than silently omit that and
    let "Gross Profit" masquerade as "Net Profit", the statement says
    exactly what it does and doesn't include, in the document itself.
    """
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Profit & Loss Statement", topMargin=3 * cm)

    title_style = ParagraphStyle("PLTitle", fontName="Helvetica-Bold", fontSize=16, spaceAfter=4)
    subtitle_style = ParagraphStyle(
        "PLSubtitle", fontName="Helvetica", fontSize=10, textColor=colors.HexColor("#475569")
    )
    note_style = ParagraphStyle(
        "PLNote",
        fontName="Helvetica-Oblique",
        fontSize=8,
        textColor=colors.HexColor("#991B1B"),
        spaceBefore=16,
    )

    def money(value: float) -> str:
        return f"{currency} {value:,.2f}"

    rows = [
        ["Revenue", money(revenue)],
        ["Cost of Goods Sold", f"({money(cost_of_goods_sold)})"],
        ["", ""],
        ["Gross Profit", money(gross_profit)],
        ["Gross Margin", f"{gross_margin_percent:.1f}%"],
    ]
    table = Table(rows, colWidths=[10 * cm, 6 * cm])
    table.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LINEABOVE", (0, 3), (-1, 3), 1, colors.HexColor("#0F172A")),
                ("FONTNAME", (0, 3), (-1, 4), "Helvetica-Bold"),
            ]
        )
    )

    elements = [
        Paragraph("Profit & Loss Statement", title_style),
        Paragraph(
            f"{business_name} &nbsp;&middot;&nbsp; {start_date} to {end_date}", subtitle_style
        ),
        Spacer(1, 16),
        table,
    ]

    if trend_points and len(trend_points) >= 2:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Revenue trend", subtitle_style))
        elements.append(Spacer(1, 6))
        elements.append(_build_trend_chart_drawing(trend_points))

    if top_products:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Revenue by product", subtitle_style))
        elements.append(Spacer(1, 6))
        elements.append(_build_product_bar_chart_drawing(top_products))

    if top_customers and len(top_customers) >= 2:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Customer revenue (Pareto)", subtitle_style))
        elements.append(Spacer(1, 6))
        elements.append(_build_customer_pareto_chart_drawing(top_customers))

    elements.append(
        Paragraph(
            "This statement reflects revenue and cost of goods sold only, computed directly "
            "from real sales and batch cost records. Operating expenses (rent, salaries, "
            "utilities, and other overhead) are not tracked anywhere in this system and are "
            "deliberately not included -- this is a Gross Profit statement, not a complete "
            "net-profit P&amp;L.",
            note_style,
        )
    )
    doc.build(elements)
    return buffer.getvalue()


def _build_trend_chart_drawing(trend_points: list["RevenueTrendPoint"]) -> "Drawing":
    """
    A real chart drawn with reportlab's own graphics support -- no
    matplotlib, which would meaningfully bloat the desktop installer
    for something shipped to many pharmacy owners on modest hardware.
    Reportlab is already a proven dependency here; this reuses it
    rather than adding a second, heavier charting stack just for PDFs.
    """
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.shapes import Drawing, String

    values = [round(p.revenue, 2) for p in trend_points]
    labels = [p.period_label for p in trend_points]
    # Reportlab's own axis renders every label -- with more than a
    # handful of points that overlaps into an unreadable smear, so
    # only a spaced-out subset gets an actual label, exactly like the
    # frontend chart's tick behavior.
    label_every = max(1, len(labels) // 8)
    sparse_labels = [lbl if i % label_every == 0 else "" for i, lbl in enumerate(labels)]

    drawing = Drawing(460, 160)
    chart = HorizontalLineChart()
    chart.x = 40
    chart.y = 20
    chart.width = 400
    chart.height = 120
    chart.data = [values]
    chart.categoryAxis.categoryNames = sparse_labels
    chart.categoryAxis.labels.fontSize = 6
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.dy = -8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labelTextFormat = "%0.0f"
    chart.lines[0].strokeColor = colors.HexColor("#8A6D3B")  # matches the app's brass accent
    chart.lines[0].strokeWidth = 1.5
    drawing.add(chart)
    if not values:
        drawing.add(String(200, 80, "No data", fontSize=9))
    return drawing


def _build_product_bar_chart_drawing(top_products: list["TopProductEntry"]) -> "Drawing":
    """
    Same reportlab-native approach as the trend chart above -- top
    products by revenue as a vertical bar chart, product names
    truncated and angled since they're often longer than a chart
    label comfortably fits.
    """
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String

    sorted_products = sorted(top_products, key=lambda p: p.revenue, reverse=True)[:8]
    values = [round(p.revenue, 2) for p in sorted_products]
    labels = [p.name[:16] for p in sorted_products]

    drawing = Drawing(460, 170)
    chart = VerticalBarChart()
    chart.x = 40
    chart.y = 40
    chart.width = 400
    chart.height = 110
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 6
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.dy = -8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labelTextFormat = "%0.0f"
    chart.bars[0].fillColor = colors.HexColor("#8A6D3B")  # matches the app's brass accent
    drawing.add(chart)
    if not values:
        drawing.add(String(200, 90, "No data", fontSize=9))
    return drawing


def _build_customer_pareto_chart_drawing(top_customers: list["TopCustomerEntry"]) -> "Drawing":
    """
    Reportlab's graphics charts don't have a built-in dual-axis combo
    type the way the web chart (bars + a cumulative-percent line) has
    -- rather than hand-building axis geometry to fake one, each bar
    is directly labeled with its own cumulative percentage, which
    reads just as clearly in a static, non-interactive PDF page.
    """
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String

    entries = top_customers[:8]
    values = [round(c.revenue, 2) for c in entries]
    labels = [c.name[:16] for c in entries]

    drawing = Drawing(460, 180)
    chart = VerticalBarChart()
    chart.x = 40
    chart.y = 40
    chart.width = 400
    chart.height = 110
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 6
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.dy = -8
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labelTextFormat = "%0.0f"
    chart.bars[0].fillColor = colors.HexColor("#8A6D3B")
    drawing.add(chart)

    if not values:
        drawing.add(String(200, 90, "No data", fontSize=9))
        return drawing

    # Cumulative-percent label directly above each bar, in the app's
    # stamp-red accent, standing in for the web chart's second line.
    max_value = max(values) or 1.0
    bar_width = 400 / len(entries)
    for i, entry in enumerate(entries):
        bar_height = (entry.revenue / max_value) * 110
        label_x = 40 + i * bar_width + bar_width / 2
        label_y = 40 + bar_height + 4
        drawing.add(
            String(
                label_x,
                label_y,
                f"{entry.cumulative_percent:.0f}%",
                fontSize=6,
                fillColor=colors.HexColor("#A13D2E"),  # matches the app's stamp-red accent
                textAnchor="middle",
            )
        )
    return drawing
