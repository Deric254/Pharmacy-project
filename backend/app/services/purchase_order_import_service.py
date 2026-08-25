"""
Purchase order bulk import -- the direct path. This creates a real
purchase order that's already RECEIVED, with real batches and real
stock movements, in one atomic step -- no draft/send/in-transit
ceremony. Matches exactly what quick_purchase() does for a single
manual entry; this is the same thing, just filled from a spreadsheet
instead of typed in one line at a time.

Same all-or-nothing guarantee as every other bulk import in this
system: every row is validated before anything is created, and if any
row has a problem, nothing is imported -- no partially-received order.
Product names must match the real, active catalog exactly (not case-
sensitive); this never invents a product from a typo'd name.
"""

import io
from datetime import date as date_type
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette.concurrency import run_in_threadpool

from app.models.product import Product
from app.models.supplier import Supplier
from app.models.user import User
from app.schemas.product import ImportRowError
from app.schemas.purchase_order import (
    PurchaseOrderOut,
    QuickPurchaseLine,
    QuickPurchaseRequest,
)
from app.services.purchasing_service import PurchasingService

_HEADERS = ["Product name", "Quantity", "Batch number", "Expiry date", "Unit cost", "Selling price"]
_EXAMPLE_ROW: list[str | int | float] = [
    "EXAMPLE - Paracetamol 500mg",
    100,
    "BATCH-001",
    "2027-06-30",
    8.5,
    12.0,
]
_MAX_ROWS = 500  # a single delivery is realistically dozens of lines, not thousands


def generate_purchase_order_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Received Stock"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    example_font = Font(name="Arial", italic=True, color="6B7280")
    for col, value in enumerate(_EXAMPLE_ROW, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = example_font

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    qty_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1=0,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid quantity",
        error="Quantity must be a whole number greater than 0.",
    )
    ws.add_data_validation(qty_validation)
    qty_validation.add(f"B2:B{_MAX_ROWS}")

    cost_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1=0,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid cost",
        error="Unit cost must be a number, 0 or greater.",
    )
    ws.add_data_validation(cost_validation)
    cost_validation.add(f"E2:E{_MAX_ROWS}")

    selling_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1=0,
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid selling price",
        error="Selling price must be a number, 0 or greater.",
    )
    ws.add_data_validation(selling_validation)
    selling_validation.add(f"F2:F{_MAX_ROWS}")

    instructions = ws.cell(
        row=1,
        column=7,
        value="Product names must match your catalog exactly. Expiry date as YYYY-MM-DD.",
    )
    instructions.font = Font(name="Arial", italic=True, size=9, color="991B1B")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _clean_str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _parse_date(value: Any) -> date_type | None:
    if isinstance(value, date_type):
        return value
    text = _clean_str(value)
    if not text:
        return None
    try:
        return date_type.fromisoformat(text)
    except ValueError:
        return None


async def _parse_and_validate(
    db: AsyncSession, file_bytes: bytes
) -> tuple[list[QuickPurchaseLine], list[ImportRowError]]:
    try:
        wb = await run_in_threadpool(load_workbook, io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="This file has no worksheet to read.")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=400, detail="Could not read this file as an Excel spreadsheet."
        ) from exc

    errors: list[ImportRowError] = []
    parsed_rows: list[tuple[int, str, int, str, date_type, float, float]] = []
    seen_batch_numbers: dict[str, int] = {}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(rows) > _MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"This file has more than {_MAX_ROWS} rows. Split it into smaller batches.",
        )

    for offset, row in enumerate(rows):
        row_num = offset + 2
        row_values: list[Any] = (list(row) + [None] * 6)[:6]
        name_raw, qty_raw, batch_raw, expiry_raw, cost_raw, selling_raw = row_values
        name = _clean_str(name_raw)

        if not name:
            continue
        if name.upper().startswith("EXAMPLE"):
            continue

        row_ok = True

        try:
            qty = int(qty_raw) if qty_raw is not None else None
            if qty is None or qty <= 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                ImportRowError(
                    row=row_num, field="Quantity", message="Must be a whole number greater than 0."
                )
            )
            row_ok = False
            qty = 0

        batch_number = _clean_str(batch_raw)
        if not batch_number:
            errors.append(
                ImportRowError(row=row_num, field="Batch number", message="Cannot be empty.")
            )
            row_ok = False
        elif batch_number in seen_batch_numbers:
            errors.append(
                ImportRowError(
                    row=row_num,
                    field="Batch number",
                    message=f"Duplicate of row {seen_batch_numbers[batch_number]} in this file.",
                )
            )
            row_ok = False
        else:
            seen_batch_numbers[batch_number] = row_num

        expiry = _parse_date(expiry_raw)
        if expiry is None:
            errors.append(
                ImportRowError(
                    row=row_num, field="Expiry date", message="Must be a real date (YYYY-MM-DD)."
                )
            )
            row_ok = False

        try:
            cost = float(cost_raw) if cost_raw is not None else None
            if cost is None or cost < 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                ImportRowError(
                    row=row_num, field="Unit cost", message="Must be a number, 0 or more."
                )
            )
            row_ok = False
            cost = 0.0

        try:
            selling_price = float(selling_raw) if selling_raw is not None else None
            if selling_price is None or selling_price < 0:
                raise ValueError
        except (TypeError, ValueError):
            errors.append(
                ImportRowError(
                    row=row_num,
                    field="Selling price",
                    message="Must be a number, 0 or more.",
                )
            )
            row_ok = False
            selling_price = 0.0

        if row_ok and expiry is not None:
            parsed_rows.append((row_num, name, qty, batch_number, expiry, cost, selling_price))

    if not parsed_rows and not errors:
        errors.append(
            ImportRowError(row=0, field="File", message="No stock rows found in this file.")
        )
        return [], errors

    # Match every candidate name against the real, active catalog in
    # one case-insensitive query -- never invents a product from a
    # typo'd name.
    names_to_match = list({name.lower() for _, name, _, _, _, _, _ in parsed_rows})
    matched_by_lower: dict[str, int] = {}
    if names_to_match:
        result = await db.execute(
            select(Product.id, func.lower(Product.name)).where(
                Product.deleted_at.is_(None), func.lower(Product.name).in_(names_to_match)
            )
        )
        matched_by_lower = dict((lower_name, pid) for pid, lower_name in result.all())

    lines: list[QuickPurchaseLine] = []
    for row_num, name, qty, batch_number, expiry, cost, selling_price in parsed_rows:
        product_id = matched_by_lower.get(name.lower())
        if product_id is None:
            errors.append(
                ImportRowError(
                    row=row_num,
                    field="Product name",
                    message="No active product with this exact name exists in your catalog.",
                )
            )
            continue
        try:
            lines.append(
                QuickPurchaseLine(
                    product_id=product_id,
                    quantity=qty,
                    batch_number=batch_number,
                    expiry_date=expiry,
                    unit_cost=cost,
                    selling_price=selling_price,
                )
            )
        except ValidationError as exc:
            errors.append(
                ImportRowError(
                    row=row_num, field="Row", message=f"Invalid data: {exc.errors()[0]['msg']}"
                )
            )

    return lines, errors


async def bulk_import_purchase_order(
    db: AsyncSession, file_bytes: bytes, supplier_id: int, user: User
) -> PurchaseOrderOut:
    supplier = await db.get(Supplier, supplier_id)
    if supplier is None:
        raise HTTPException(status_code=404, detail="Supplier not found")

    lines, errors = await _parse_and_validate(db, file_bytes)

    if errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": f"{len(errors)} problem(s) found. Nothing was received.",
                "errors": [e.model_dump() for e in errors],
            },
        )

    payload = QuickPurchaseRequest(supplier_id=supplier_id, lines=lines)
    return await PurchasingService(db).quick_purchase(payload, user)
